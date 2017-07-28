from openpyxl.styles import Font
from openpyxl import Workbook
from utils.tool_funcs import write_df_to_excel, tradecode_to_windcode
import os
import pandas as pd


def save_details_to_excel(factor,file_path, env):
    """保存的内容主要包括：
    sheet1:第一组表现
    sheet2：多空组合表现
    sheet3：分组信息
    """
    nrows = 1 #记录已占用行数
    wb = Workbook()
    # 第一个sheet
    active_sheet = wb.active
    active_sheet.title = "第一组表现"
    ft = Font(bold=True)
    
    start_point = (1,1)
    for _m in factor.first_group_active_return.long_short_returns:
        first_group_pf = factor.first_group_active_return.long_short_returns[_m]
        yearly_performance = first_group_pf.YearlyPerformance()
        active_sheet.merge_cells(start_row=start_point[0],start_column=1,end_row=start_point[0],end_column=5)
        _ = active_sheet.cell(row=start_point[0],column=1,value="第一组分年表现(超额收益,%s)"%_m)
        start_point = (start_point[0] + 1, start_point[1])
        end_point = write_df_to_excel(active_sheet,start_point,yearly_performance)
        start_point = (end_point[0]+2,1)
    
    # 在sheet1的右侧同时记录第一组的日收益率
    _list = []
    for _m in factor.first_group_active_return.long_short_returns:
        _list.append(
            pd.Series(factor.first_group_active_return.long_short_returns[_m].activeRet,name=_m))
    df = pd.concat(_list,axis=1)
    df_shape = df.shape
    active_sheet.merge_cells(start_row=1,start_column=8,end_row=1,end_column=8+df_shape[1])
    _ = active_sheet.cell(
        row=1,column=8,value="第一组超额收益率(基准:{benchmark})".format(benchmark=env._config.benchmark))
    write_df_to_excel(active_sheet,(2,8),df)
    
    # 在sheet1的右侧记录第一组的月收益率
    _list = []
    for _m in factor.first_group_active_return.long_short_returns:
        _list.append(
            pd.Series(factor.first_group_active_return.long_short_returns[_m].MonthlyRet(),name=_m))
    df = pd.concat(_list,axis=1)
    df_shape = df.shape
    active_sheet.merge_cells(start_row=1,start_column=13,end_row=1,end_column=13+df_shape[1])
    _ = active_sheet.cell(
        row=1,column=13,value="第一组月超额收益率(基准:{benchmark})".format(benchmark=env._config.benchmark))
    write_df_to_excel(active_sheet, (2, 13), df)    
    
    # 第二个sheet
    start_point = (1,1)
    active_sheet = wb.create_sheet("多空组合表现")
    for _m in factor.long_short_return.long_short_returns:
        long_short_pf = factor.long_short_return.long_short_returns[_m]
        yearly_performance = long_short_pf.YearlyPerformance()
        active_sheet.merge_cells(start_row=start_point[0],start_column=1,end_row=start_point[0],end_column=5)
        _ = active_sheet.cell(row=start_point[0],column=1,value="多空分组分年表现(%s)"%_m)
        start_point = (start_point[0] + 1, start_point[1])
        end_point = write_df_to_excel(active_sheet,start_point,yearly_performance)
        start_point = (end_point[0]+2,1)
    
    # 在sheet2的右侧同时记录多空组合的日收益率
    _list = []
    for _m in factor.long_short_return.long_short_returns:
        _list.append(
            pd.Series(factor.long_short_return.long_short_returns[_m].activeRet,name=_m))
    df = pd.concat(_list,axis=1)
    df_shape = df.shape
    active_sheet.merge_cells(start_row=1,start_column=8,end_row=1,end_column=8+df_shape[1])
    _ = active_sheet.cell(row=1,column=8,value="多空组合收益率")
    write_df_to_excel(active_sheet,(2,8), df)
    
    # 在sheet2的右侧记录多空月收益率
    _list = []
    for _m in factor.long_short_return.long_short_returns:
        _list.append(
            pd.Series(factor.long_short_return.long_short_returns[_m].MonthlyRet(),name=_m))
    df = pd.concat(_list,axis=1)
    df_shape = df.shape
    active_sheet.merge_cells(start_row=1,start_column=13,end_row=1,end_column=13+df_shape[1])
    _ = active_sheet.cell(
        row=1,column=13,value="多空月超额收益率")
    write_df_to_excel(active_sheet, (2, 13), df)       
    
    # 在第三个sheet中记录IC序列
    start_point = (1, 1)
    active_sheet = wb.create_sheet("IC序列")
    ic = factor.ic_series.to_frame()
    ic.columns = ['IC']
    ic.index.name = 'date'
    write_df_to_excel(active_sheet, start_point, ic)

    # 保存excel
    file_name = file_path + os.sep+ factor.name + '_details.xlsx'
    wb.save(filename=file_name)

def save_summary_to_excel(factors, file_path, env, method='typical'):
    """按照兴业证券的简要模板存储"""
    factor_names = []
    ic_series = []
    long_short = []
    long_only = []
    for factor in factors:
        factor_names.append(factor.name)
        ic_series.append(factor.ic_series.describe())
        long_only.append(factor.first_group_active_return.long_short_returns[method].get_summary())
        long_short.append(factor.long_short_return.long_short_returns[method].get_summary())
    factor_name_dict = {x: factor_names[x] for x in range(len(factor_names))}
    ic_series = pd.concat(ic_series, axis=1).rename_axis(factor_name_dict, axis='columns').T
    long_only = pd.concat(long_only, axis=1).rename_axis(factor_name_dict, axis='columns').T
    long_short = pd.concat(long_short, axis=1).rename_axis(factor_name_dict, axis='columns').T
    
    wb = Workbook()
    active_sheet = wb.active
    active_sheet.title = "IC"
    write_df_to_excel(active_sheet, (1, 1), ic_series)
    active_sheet = wb.create_sheet("分位数多头组合")
    write_df_to_excel(active_sheet, (1, 1), long_only)
    active_sheet = wb.create_sheet("分位数多空组合")
    write_df_to_excel(active_sheet, (1, 1), long_short)
    file_name = file_path + os.sep + 'summary.xlsx'
    wb.save(filename=file_name)
    
def save_stock_list(factor, file_path, env):
    """
    存储股票列表
    """
    for _m in factor.stock_list:
        stock_list = factor.stock_list[_m].reset_index()
        stock_list['IDs'] = stock_list['IDs'].apply(tradecode_to_windcode)
        file_name = "_".join([os.path.join(file_path,factor.name), 'stock_list', _m+'.csv'])
        stock_list.to_csv(file_name, index=False)
    

