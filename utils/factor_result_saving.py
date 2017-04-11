from openpyxl.styles import Alignment, Font
from openpyxl import Workbook
from utils.tool_funcs import write_df_to_excel, tradecode_to_windcode
from utils.datetime_func import Datetime2IntDate
from factor_performance.fund_performance import FundPerformance
from scipy.io import savemat
import os
import pandas as pd
from contextlib import closing
import shelve

def save_ic_figure(factor, file_path,config,env):
    file_name = file_path + os.sep+ factor.name + '_ic_series.png'
    factor.ic_fig.savefig(file_name)

def save_group_return_figure(factor, file_path,config,env):
    file_name = file_path + os.sep+ factor.name + '_group_returns.png'
    factor.group_return_fig.savefig(file_name)
 
def save_details_to_excel(factor,file_path,config,env):
    """保存的内容主要包括：
    sheet1:第一组表现
    sheet2：多空组合表现
    sheet3：分组信息
    """
    nrows = 1 #记录已占用行数
    wb = Workbook()
    # 第一个sheet
    active_sheet = wb.active
    active_sheet.title = "第一组表现"
    ft = Font(bold=True)
    
    start_point = (1,1)
    for _m in factor.first_group_active_return:
        first_group_pf = factor.first_group_active_return[_m]
        yearly_performance = first_group_pf.YearlyPerformance()
        active_sheet.merge_cells(start_row=start_point[0],start_column=1,end_row=start_point[0],end_column=5)
        _ = active_sheet.cell(row=start_point[0],column=1,value="第一组分年表现(超额收益,%s)"%_m)
        start_point = (start_point[0] + 1, start_point[1])
        end_point = write_df_to_excel(active_sheet,start_point,yearly_performance)
        start_point = (end_point[0]+2,1)
    
    # 在sheet1的右侧同时记录第一组的日收益率
    _list = []
    for _m in factor.first_group_active_return:
        _list.append(pd.Series(factor.first_group_active_return[_m].activeRet,name=_m))
    df = pd.concat(_list,axis=1)
    df_shape = df.shape
    active_sheet.merge_cells(start_row=1,start_column=8,end_row=1,end_column=8+df_shape[1])
    _ = active_sheet.cell(row=1,column=8,value="第一组超额收益率(基准:{benchmark})".format(benchmark=config.benchmark))
    write_df_to_excel(active_sheet,(2,8),df)
    
    # 第二个sheet
    start_point = (1,1)
    active_sheet = wb.create_sheet("多空组合表现")
    for _m in factor.long_short_return:
        long_short_pf = factor.long_short_return[_m]
        yearly_performance = long_short_pf.YearlyPerformance()
        active_sheet.merge_cells(start_row=start_point[0],start_column=1,end_row=start_point[0],end_column=5)
        _ = active_sheet.cell(row=start_point[0],column=1,value="多空分组分年表现(%s)"%_m)
        start_point = (start_point[0] + 1, start_point[1])
        end_point = write_df_to_excel(active_sheet,start_point,yearly_performance)
        start_point = (end_point[0]+2,1)
    
    # 在sheet2的右侧同时记录多空组合的日收益率
    _list = []
    for _m in factor.long_short_return:
        _list.append(pd.Series(factor.long_short_return[_m].activeRet,name=_m))
    df = pd.concat(_list,axis=1)
    df_shape = df.shape
    active_sheet.merge_cells(start_row=1,start_column=8,end_row=1,end_column=8+df_shape[1])
    _ = active_sheet.cell(row=1,column=8,value="多空组合收益率")
    write_df_to_excel(active_sheet,(2,8), df)
    
    # 在第三个sheet中记录IC序列
    start_point = (1, 1)
    active_sheet = wb.create_sheet("IC序列")
    ic = factor.ic_series.to_frame()
    ic.columns = ['IC']
    ic.index.name = 'date'
    write_df_to_excel(active_sheet, start_point, ic)
    
    
    # 保存excel
    file_name = file_path + os.sep+ factor.name + '_details.xlsx'
    wb.save(filename=file_name)

def save_factor_as_pickle(factor, file_path, config, env):
    """
    将因子的信息保存到shelve中
    """
    file_name = file_path + os.sep + factor.name + "_data"
    with closing(shelve.open(file_name)) as f:
        f['group_return'] = factor.group_return                             # 保存因子的分组收益率
        f['group_info'] = factor.grouping_info                              # 保存因子的分组明细
        f['ic_series'] = factor.ic_series                                   # 保存因子的IC序列
        f['first_group_active_return'] = factor.first_group_active_return   # 保存因子第一组的超额收益率
        f['factor_data'] = factor.data                                      # 保存因子数据
        f['config'] = config
        f['env'] = env
        f['factor_tuple'] = (factor.name, factor.axe, factor.direction)

def save_group_info_as_mat(factor, file_path, config, env):
    """
    把因子的分组信息保存成matlab文件格式：
    1.股票代码 2.日期 3 模型1分组ID 4 模型2分组ID 5 模型3分组ID
    """
    _l = []
    for _m in factor.grouping_info:
        _l.append(factor.grouping_info[_m])
    group_info = pd.concat(_l, axis=1, ignore_index=True)
    group_info.columns = list(factor.grouping_info)
    group_info.reset_index(inplace=True)
    group_info['IDs'] = group_info['IDs'].apply(tradecode_to_windcode)
    group_info['date'] = group_info['date'].apply(Datetime2IntDate)
    
    data = group_info.values.astype('O')
    column = group_info.columns.values
    data_dict = {'group_info': data, 'column_name': column}
    
    file_name = file_path + os.sep + factor.name + '_group_info.mat'
    savemat(file_name, data_dict)

    
    